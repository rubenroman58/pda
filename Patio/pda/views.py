from django.shortcuts import render, redirect
from datetime import datetime,timedelta,date
from django.db.models import Sum,Q
from django.shortcuts import get_object_or_404
from .models import Patio, Paquete,AlbaranDevolucion,LineaArticulo,TipoTarea,Trabajador,Articulo
from .forms import PatioForm,PaqueteForm,AlbaranForm,LineaArticulo,LineaArticuloForm,TrabajadorForm
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import TemplateView
from openpyxl import Workbook
from django.http import HttpResponse
from .utils import get_articulos_dict,get_tipos_tarea_dict,get_trabajadores_dict
from collections import defaultdict
from pda.models import Articulo, Andalucia, Levante, Madrid, Cataluña
import pandas as pd

from openpyxl import load_workbook

from django.conf import settings

from io import BytesIO

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from openpyxl.utils import get_column_letter



def iniciar_tarea(request):
    if request.method == 'POST':
        form = PatioForm(request.POST)
        if form.is_valid():
            tarea = form.save(commit=False)
            tarea.fecha = datetime.now().date()
            tarea.horaInicio = datetime.now().time().replace(microsecond=0)  
            tarea.save()  

            return redirect('crear_paquete', tarea_id=tarea.id)
    else:
        form = PatioForm()

    trabajadores_dict = get_trabajadores_dict()
    tipos_tarea_dict = get_tipos_tarea_dict()

    return render(request, 'iniciar_tarea.html', {
        'form': form,
        'trabajadores': trabajadores_dict,
        'tipos_tarea': tipos_tarea_dict,
    })


def crear_paquete(request, tarea_id):
    tarea = get_object_or_404(Patio, id=tarea_id)

    if request.method == 'POST':
        form = PaqueteForm(request.POST)
        if form.is_valid():
            ahora = datetime.now().time().replace(microsecond=0)
            
            # Si es el primer paquete, usamos la hora de inicio de la tarea
            if not Paquete.objects.filter(tarea=tarea).exists():
                hora_inicio_paquete = tarea.horaInicio  # Para el primer paquete, la hora es la de la tarea
                hora_fin_paquete = ahora  # Hora de fin del primer paquete, la hora actual
            else:
                # Para los paquetes siguientes, la hora de inicio es la hora de fin del último paquete
                ultimo_paquete = Paquete.objects.filter(tarea=tarea).order_by('-id').first()
                hora_inicio_paquete = ultimo_paquete.horaFin  # La hora de inicio será la hora de fin del último paquete
                hora_fin_paquete = ahora  # La hora de fin será la hora actual

            # Crea el paquete sin guardarlo todavía
            paquete = form.save(commit=False)
            paquete.tarea = tarea
            paquete.horaInicio = hora_inicio_paquete  # Asignamos la hora de inicio al paquete
            paquete.horaFin = hora_fin_paquete  # Asignamos la hora de fin al paquete

            # Guardamos el paquete
            paquete.save()
            print(f"Nuevo paquete creado: horaInicio = {paquete.horaInicio}, horaFin = {paquete.horaFin}")

            # Si ya existía un paquete, actualizamos la hora de fin del paquete anterior
            if Paquete.objects.filter(tarea=tarea).count() > 1:
                ultimo_paquete = Paquete.objects.filter(tarea=tarea).order_by('-id').first()
                ultimo_paquete.horaFin = hora_fin_paquete  # El último paquete termina cuando empieza el nuevo
                ultimo_paquete.save()
                print(f"Hora fin del paquete anterior (ID {ultimo_paquete.id}) actualizada a {hora_fin_paquete}")

            # Actualizar la cantidad total
            total_cantidad = Paquete.objects.filter(tarea=tarea).aggregate(Sum('cantidad_paquete'))['cantidad_paquete__sum'] or 0
            tarea.cantidad = total_cantidad
            tarea.save()

            # Redirige a la misma página de crear paquete o a donde desees
            return redirect('crear_paquete', tarea_id=tarea.id)

        else:
            # Si el formulario no es válido, asegúrate de devolver una respuesta
            # En este caso, simplemente renderizamos el formulario con los errores
            articulos_dict = get_articulos_dict()
            return render(request, 'crear_paquete.html', {
                'form': form,
                'tarea': tarea,
                'articulos': articulos_dict,
            })
    else:
        form = PaqueteForm()
        articulos_dict = get_articulos_dict()
        return render(request, 'crear_paquete.html', {
            'form': form,
            'tarea': tarea,
            'articulos': articulos_dict,
        })


def finalizar_tarea(request, tarea_id):
    tarea = get_object_or_404(Patio, id=tarea_id)
    tarea.horaFin = datetime.now().time().replace(microsecond=0)  
    tarea.save()
    return redirect('iniciar_tarea') 


def seleccionar_albaran(request):
    if request.method == 'POST':
        form = AlbaranForm(request.POST)
        if form.is_valid():
            numero = form.cleaned_data['numero']
            albaran, creado = AlbaranDevolucion.objects.get_or_create(numero=numero)
          
            return redirect('agregar_lineas', albaran_id=albaran.id)
        else:
            print("Este formulario ya ha sido creado:", form.errors)
    else:
        form = AlbaranForm()
    return render(request, 'seleccionar_albaran.html', {'form': form})


class HomeView (TemplateView): 
    template_name='index.html'


def agregar_lineas(request,albaran_id):
    albaran = AlbaranDevolucion.objects.get(id=albaran_id)
    if request.method=='POST':
        form=LineaArticuloForm(request.POST)
        if form.is_valid():
         cantidad_buena= form.cleaned_data['cantidad_buena']
         cantidad_mala= form.cleaned_data['cantidad_mala']
         chatarra= form.cleaned_data['chatarra']
         idArticulo= form.cleaned_data['idArticulo']

         LineaArticulo.objects.create(
             albaran=albaran,
             idArticulo=idArticulo,
             cantidad_buena=cantidad_buena,
             chatarra=chatarra,
             cantidad_mala=cantidad_mala
         )
         return redirect('agregar_lineas', albaran_id=albaran.id)
        else:           
            return render(request,'agregar_lineas.html',{
                'form':form,
                'albaran':albaran,
                'articulos':get_articulos_dict(),
            })
    else:
        form = LineaArticuloForm()
        articulos_dict = get_articulos_dict()
        return render(request, 'agregar_lineas.html', {
            'form': form, 
            'albaran': albaran,
            'articulos':articulos_dict
             })
    

def agregar_lineas2(request,albaran_id):
    albaran = AlbaranDevolucion.objects.get(id=albaran_id)
    if request.method=='POST':
        form=LineaArticuloForm(request.POST)
        if form.is_valid():
         cantidad_buena= form.cleaned_data['cantidad_buena']
         cantidad_mala= form.cleaned_data['cantidad_mala']
         chatarra= form.cleaned_data['chatarra']
         idArticulo= form.cleaned_data['idArticulo']

         LineaArticulo.objects.create(
             albaran=albaran,
             idArticulo=idArticulo,
             cantidad_buena=cantidad_buena,
             chatarra=chatarra,
             cantidad_mala=cantidad_mala
         )
         return redirect('añadir_lineas', albaran_id=albaran.id)
        else:           
            return render(request,'añadir_linea.html',{
                'form':form,
                'albaran':albaran,
                'articulos':get_articulos_dict(),
            })
    else:
        form = LineaArticuloForm()
        articulos_dict = get_articulos_dict()
        return render(request, 'añadir_linea.html', {
            'form': form, 
            'albaran': albaran,
            'articulos':articulos_dict
             })


def salir(request):
    return render(request,'cerrar_programa.html')


def estadisticas(request):

    return render(request,'estadisticas.html')


def lista_trabajadores(request):
    trabajadores = Trabajador.objects.all()
    return render(request, 'lista_trabajadores.html', {'trabajadores_lista': trabajadores})


def lista_tareas_completa(request):
    tareas = Patio.objects.all().order_by('-fecha')
    
    tareas_info = []
    for tarea in tareas:
        tarea_info = {
            'id': tarea.id,
            'fecha': tarea.fecha,
            'horaInicio': tarea.horaInicio,
            'horaFin': tarea.horaFin,
            'tipo_tarea': TipoTarea.objects.filter(id=tarea.idTipTarea).first(),
            'operador1': Trabajador.objects.filter(id=tarea.idOper1).first(),
            'operador2': Trabajador.objects.filter(id=tarea.idOper2).first() if tarea.idOper2 else None,
            'cantidad': tarea.cantidad
        }
        tareas_info.append(tarea_info)

    return render(request, 'lista_tareas.html', {
        'tareas_info': tareas_info
    })


def detalles_tarea(request,tarea_id):
    tarea=get_object_or_404(Patio,id=tarea_id)
    paquetes=Paquete.objects.filter(tarea=tarea)
    volver_url = request.META.get('HTTP_REFERER', '/lista_tareas/')
    return render(request,'detalle_tarea.html',{
        'tarea':tarea,
        'paquetes':paquetes,
        'volver_url':volver_url
        })


def estadisticas_trabajador(request, trabajador_id):
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    tareas = Patio.objects.filter(Q(idOper1=trabajador_id) | Q(idOper2=trabajador_id))
    
    # Obtener el periodo de la query string (día, semana, mes, todo)
    periodo = request.GET.get('periodo')
    hoy = date.today()
    
    # Variables generales
    cantidadTotal = 0
    tiempoTotalSegundos = 0
    productividad = 0
    numTareas = 0
    tiempoPromedio = "00:00:00"
    
    # Diccionario para almacenar estadísticas por tipo de tarea
    estadisticas_por_tipo = defaultdict(lambda: {
        'cantidad': 0,
        'tiempo_total': 0,
        'num_tareas': 0,
        'tareas': []
    })
    
    # Filtrar tareas según el periodo
    if periodo == 'dia':
        tareas = tareas.filter(fecha=hoy)
    elif periodo == 'semana':
        inicio_semana = hoy - timedelta(days=hoy.weekday())  # Primer día de la semana
        tareas = tareas.filter(fecha__gte=inicio_semana, fecha__lte=hoy)
    elif periodo == 'mes':
        tareas = tareas.filter(fecha__month=hoy.month, fecha__year=hoy.year)
    elif periodo == 'todo':
        pass  # No filtra, usa todas las tareas

    # Agrupar las tareas por tipo
    for tarea in tareas:
        if tarea.horaInicio and tarea.horaFin and tarea.idTipTarea is not None:
            try:
                tipo_tarea = TipoTarea.objects.get(id=tarea.idTipTarea)
                nombre_tarea = tipo_tarea.nombre
            except TipoTarea.DoesNotExist:
                nombre_tarea = f"Tarea ID {tarea.idTipTarea}"

            # Calcular tiempo de la tarea
            fecha = tarea.fecha
            hora_inicio = datetime.combine(fecha, tarea.horaInicio)
            hora_fin = datetime.combine(fecha, tarea.horaFin)
            tiempo_segundos = (hora_fin - hora_inicio).total_seconds()

            # Acumular estadísticas por tipo de tarea
            estadisticas_por_tipo[nombre_tarea]['cantidad'] += tarea.cantidad or 0
            estadisticas_por_tipo[nombre_tarea]['tiempo_total'] += tiempo_segundos
            estadisticas_por_tipo[nombre_tarea]['num_tareas'] += 1
            estadisticas_por_tipo[nombre_tarea]['tareas'].append(tarea)

    # Calcular productividad y tiempo promedio
    for nombre_tarea, datos in estadisticas_por_tipo.items():
        tiempo_total = datos['tiempo_total']
        num_tareas = datos['num_tareas']
        cantidad = datos['cantidad']

        # Calcular productividad (cantidad por hora)
        if tiempo_total > 0:
            datos['productividad'] = round(cantidad / (tiempo_total / 3600), 2)
        else:
            datos['productividad'] = 0

        # Calcular tiempo promedio por tarea
        if num_tareas > 0:
            tiempo_promedio = tiempo_total / num_tareas
            h = int(tiempo_promedio // 3600)
            m = int((tiempo_promedio % 3600) // 60)
            s = int(tiempo_promedio % 60)
            datos['tiempo_promedio'] = f"{h:02d}:{m:02d}:{s:02d}"
        else:
            datos['tiempo_promedio'] = "00:00:00"

    # Preparar los datos para el template
    return render(request, 'estadisticas_trabajador.html', {
        'trabajador': trabajador,
        'tareas': tareas,
        'estadisticas_por_tipo': dict(estadisticas_por_tipo),
        'periodo': periodo,
    })


def lista_albaranes_completa(request):

    listaAlbaranes=AlbaranDevolucion.objects.all().order_by('-fecha')
    albaranesinfo=[]
    for albaran in listaAlbaranes:
        albaraninfo={
        'id':albaran.id,
        'numero': albaran.numero,
        'fecha':albaran.fecha
        }
        albaranesinfo.append(albaraninfo)
    return render (request,'lista_albaranes.html',{
        'albaranesinfo':albaranesinfo
    }) 


def detalles_albaran(request,albaran_id):
    albaran=get_object_or_404(AlbaranDevolucion,id=albaran_id)
    articulos=LineaArticulo.objects.filter(albaran=albaran) #Obtenemos las lineas asociadas al albaran
    articulos_dict = get_articulos_dict()
    for linea in articulos:
        linea.nombre_articulo=articulos_dict.get(linea.idArticulo)
    return render(request,'detalle_albaran.html',{
        'albaran':albaran,
        'articulos':articulos,
    })


def informacionIndek(request):

    return render(request,'informacionIndek.html')


def eliminar_alabarn(request, albaran_id):
    albaran = get_object_or_404(AlbaranDevolucion, id=albaran_id)
    albaran.delete()
    return redirect('listaAlbaranes')


def eliminar_tarea(request,tarea_id):
    tarea=get_object_or_404(Patio,id=tarea_id)
    tarea.delete()
    return redirect('listaTareas')

@csrf_exempt 
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            print('Login correcto')
            return redirect('home')  # Esto depende de que la URL 'index' esté definida en tu URLconf
        else:
            print('Login fallido')
            messages.error(request, 'Usuario o contraseña incorrecta')

    return render(request, 'login.html')
 
    
def editar_linea_articulo(request, linea_id):
    linea = get_object_or_404(LineaArticulo, id=linea_id)
    albaran = linea.albaran

    if request.method == 'POST':
        form = LineaArticuloForm(request.POST, instance=linea)
        if form.is_valid():
            form.save()
            return redirect('detalle_albaran', albaran_id=albaran.id)
    else:
        form = LineaArticuloForm(instance=linea)

    return render(request, 'editar_linea.html', {
        'form': form,
        'albaran': albaran
    })


def eliminar_linea_articulo(request, linea_id):
    linea = get_object_or_404(LineaArticulo, id=linea_id)
    albaran_id = linea.albaran.id
    linea.delete()
    return redirect('detalle_albaran', albaran_id=albaran_id)


def exportar_trabajadores_excel(request):
   
   wb=Workbook()
   ws=wb.active
   ws.title='Trabajadores'
   ws.append(['Id','Nombre'])

   for trabajador in Trabajador.objects.all():
       ws.append([trabajador.id,trabajador.nombre])


   response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
   response['Content-Disposition'] = 'attachment; filename=trabajadores.xlsx'
   wb.save(response)
   return response


def comparativa_productividad(request):
    

    periodo = request.GET.get('periodo')
    hoy = date.today()

    # Obtener tareas según el periodo
    tareas = Patio.objects.all()
    if periodo == 'dia':
        tareas = tareas.filter(fecha=hoy)
    elif periodo == 'semana':
        inicio_semana = hoy - timedelta(days=hoy.weekday())
        tareas = tareas.filter(fecha__gte=inicio_semana, fecha__lte=hoy)
    elif periodo == 'mes':
        tareas = tareas.filter(fecha__month=hoy.month, fecha__year=hoy.year)

    # Diccionario para estadísticas
    estadisticas_por_trabajador = defaultdict(lambda: {
        'cantidad': 0,
        'tiempo_total': 0,
        'num_tareas': 0,
        'productividad': 0,
        'nombre': '',
    })

    for tarea in tareas:
        if tarea.horaInicio and tarea.horaFin:
            hora_inicio = datetime.combine(tarea.fecha, tarea.horaInicio)
            hora_fin = datetime.combine(tarea.fecha, tarea.horaFin)
            tiempo_segundos = (hora_fin - hora_inicio).total_seconds()

            # Obtener los trabajadores por su ID
            for operador_id in [tarea.idOper1, tarea.idOper2]:
                if operador_id:
                    # Buscar el trabajador correspondiente
                    try:
                        trabajador_obj = Trabajador.objects.get(id=operador_id)
                        stats = estadisticas_por_trabajador[operador_id]
                        stats['nombre'] = trabajador_obj.nombre
                        stats['cantidad'] += tarea.cantidad or 0
                        stats['tiempo_total'] += tiempo_segundos
                        stats['num_tareas'] += 1
                    except Trabajador.DoesNotExist:
                        # Si no se encuentra el trabajador
                        stats = estadisticas_por_trabajador[operador_id]
                        stats['nombre'] = f'Operador ID {operador_id} no encontrado'

    # Calcular productividad
    for stats in estadisticas_por_trabajador.values():
        if stats['tiempo_total'] > 0:
           stats['productividad'] = round(stats['cantidad'] / (stats['tiempo_total'] / 3600), 2)

    # Ordenar trabajadores por productividad
    trabajadores_ordenados = sorted(estadisticas_por_trabajador.items(), key=lambda x: x[1]['productividad'], reverse=True)

    return render(request, 'comparativa_productividad.html', {
        'trabajadores': trabajadores_ordenados,
        'periodo': periodo
    })
    
def exportar_datos(request):
    delegaciones = ['Andalucia', 'Levante', 'Madrid', 'Cataluña']
    modelos = {
        'Andalucia': Andalucia,
        'Levante': Levante,
        'Madrid': Madrid,
        'Cataluña': Cataluña
    }

    resultados = []
    columnas = ['Articulo', 'Nombre']
    for deleg in delegaciones:
        columnas.extend([
            f'{deleg} Tot.Fact.Alq.Dia', f'{deleg} Tot.Unid',
            f'{deleg} P.Alq.Medio', f'{deleg} %Fact'
        ])
    columnas.extend([
        'General Tot.Fact.Alq.Dia', 'General Tot.Unid',
        'General P.Alq.Medio', 'General %Fact'
    ])

    # 🔹 Total facturación por delegación (para calcular %Fact)
    total_fact_deleg = {}
    for deleg in delegaciones:
        total = 0
        for data in modelos[deleg].objects.all():
            total += data.tot_unid * data.p_alq_medio
        total_fact_deleg[deleg] = total / 100

    totales_por_deleg = {deleg: 0 for deleg in delegaciones}

    for articulo in Articulo.objects.all():
        if not any(modelos[deleg].objects.filter(articulo=articulo).exists() for deleg in delegaciones):
            continue

        fila = [articulo.id, articulo.nombre]
        general_total_fact = 0
        general_total_unid = 0

        # Obtener datos por delegación y acumular generales
        data_por_deleg = {}
        for deleg in delegaciones:
            data = modelos[deleg].objects.filter(articulo=articulo).first()
            data_por_deleg[deleg] = data
            if data:
                tot = data.tot_unid * data.p_alq_medio / 100
                totales_por_deleg[deleg] += tot
                general_total_fact += tot
                general_total_unid += data.tot_unid

        # Rellenar fila por delegación
        for deleg in delegaciones:
            data = data_por_deleg[deleg]
            if data:
                tot = data.tot_unid * data.p_alq_medio / 100
                p_medio = data.p_alq_medio / 100
                total_deleg = total_fact_deleg[deleg]
                porcentaje = (tot / total_deleg * 100) if total_deleg else 0
                fila += [
                    f'{tot:,.2f}', f'{data.tot_unid:,}', f'{p_medio:.4f}', f'{porcentaje:.2f}%'
                ]
            else:
                fila += ['-', '-', '-', '-']

        # General
        if general_total_unid:
            p_general_medio = general_total_fact / general_total_unid
        else:
            p_general_medio = 0
        fila += [
            f'{general_total_fact:,.2f}', f'{general_total_unid:,}',
            f'{p_general_medio:.4f}', '100.00%'
        ]

        resultados.append(fila)

    df = pd.DataFrame(resultados)

    output = BytesIO()
    df.to_excel(output, index=False, header=False, startrow=4)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    # Estilos
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    colores = {
        'Andalucia': 'FFF2CC', 'Levante': 'D9EAD3',
        'Madrid': 'CFE2F3', 'Cataluña': 'F4CCCC'
    }

    # Título
    fecha = datetime.today().strftime('%d/%m/%Y')
    titulo = f"Ranking Articulos Comparativo Obras Activas por Delegación a Fecha: {fecha}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    c = ws.cell(row=1, column=1)
    c.value = titulo
    c.font = Font(size=14, bold=True, color='FFFFFF')
    c.alignment = center
    c.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    # Encabezados
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.cell(row=2, column=1, value='Articulo').font = bold
    ws.cell(row=2, column=1).alignment = center

    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    ws.cell(row=2, column=2, value='Nombre').font = bold
    ws.cell(row=2, column=2).alignment = center

    start_col = 3
    for deleg in delegaciones:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 3)
        cell = ws.cell(row=2, column=start_col)
        cell.value = deleg
        cell.font = bold
        cell.alignment = center
        fill = PatternFill(start_color=colores[deleg], end_color=colores[deleg], fill_type='solid')
        for col in range(start_col, start_col + 4):
            ws.cell(row=2, column=col).fill = fill
            ws.cell(row=3, column=col).fill = fill
        start_col += 4

    # General
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 3)
    cell = ws.cell(row=2, column=start_col)
    cell.value = 'General'
    cell.font = bold
    cell.alignment = center
    for col in range(start_col, start_col + 4):
        ws.cell(row=3, column=col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    # Subencabezados
    subcols = ['Tot.Fact.Alq.Dia', 'Tot.Unid', 'P.Alq.Medio', '%Fact']
    col = 3
    for _ in delegaciones + ['General']:
        for sub in subcols:
            cell = ws.cell(row=3, column=col)
            cell.value = sub
            cell.font = bold
            cell.alignment = center
            cell.border = border
            col += 1

    # Totales de alquiler por día
    fila_total = ['-', 'TOTAL ALQUILER POR DÍA']
    for deleg in delegaciones:
        total = totales_por_deleg[deleg]
        fila_total += [f'{total:,.2f}', '-', '-', '-']
    fila_total += ['-', '-', '-', '-']
    ws.append(fila_total)

    # Ajuste de anchos
    for i, col_cells in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
        ws.column_dimensions[get_column_letter(i)].width = max_len + 2
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    response = HttpResponse(final_output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="ranking_articulos_delegaciones.xlsx"'
    return response
